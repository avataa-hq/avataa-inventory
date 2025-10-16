"""Tests for batch router"""

import json
import pickle
import random
from datetime import datetime
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session, select

from models import TMO, TPRM, MO, PRM
from test_routers.test_batch_import import generate_csv_in_memory

URL = "/api/inventory/v1/batch/batch_objects_preview/"

TMO_0_DEFAULT_DATA = {
    "name": "Test TMO 1",
    "version": 1,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TMO_DEFAULT_DATA = {
    "name": "Test TMO",
    "version": 1,
    "p_id": 1,
    "primary": [1, 2],
    "created_by": "Test admin",
    "modified_by": "Test admin",
    "global_uniqueness": False,
    "points_constraint_by_tmo": [1, 2],
}

TPRM_1_STR_DEFAULT_DATA = {
    "name": "Test str TPRM",
    "val_type": "str",
    "required": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}
TPRM_2_STR_DEFAULT_DATA = {
    "name": "Test str_1 TPRM",
    "val_type": "str",
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_2_INT_DEFAULT_DATA = {
    "name": "Test int TPRM",
    "val_type": "int",
    "required": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_3_BOOL_DEFAULT_DATA = {
    "name": "Test bool TPRM",
    "tmo_id": 1,
    "val_type": "bool",
    "required": False,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_4_DATE_DEFAULT_DATA = {
    "name": "Test date TPRM",
    "tmo_id": 1,
    "val_type": "date",
    "required": False,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_5_DATETIME_DEFAULT_DATA = {
    "name": "Test datetime TPRM",
    "tmo_id": 1,
    "val_type": "datetime",
    "required": False,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_6_MULTI_INT_DEFAULT_DATA = {
    "name": "Test multiple int TPRM",
    "tmo_id": 1,
    "val_type": "int",
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_7_MULTI_STR_DEFAULT_DATA = {
    "name": "Test multiple str TPRM",
    "tmo_id": 1,
    "val_type": "str",
    "constraint": "^[a-zA-Z]+$",
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_8_MULTI_FLOAT_DEFAULT_DATA = {
    "name": "Test multiple float TPRM",
    "tmo_id": 1,
    "val_type": "float",
    "required": False,
    "multiple": True,
    # 'constraint': '0.0:100.0',
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_9_MULTI_DATE_DEFAULT_DATA = {
    "name": "Test multiple date TPRM",
    "tmo_id": 1,
    "val_type": "date",
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_10_MULTI_DATETIME_DEFAULT_DATA = {
    "name": "Test multiple datetime TPRM",
    "tmo_id": 1,
    "val_type": "datetime",
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_11_MULTI_BOOL_DEFAULT_DATA = {
    "name": "Test multiple bool TPRM",
    "tmo_id": 1,
    "val_type": "bool",
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_12_MULTI_MO_LINK_DEFAULT_DATA = {
    "name": "Test multiple mo_link TPRM",
    "tmo_id": 1,
    "val_type": "mo_link",
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_13_MULTI_PRM_LINK_DEFAULT_DATA = {
    "name": "Test multiple prm_link TPRM",
    "tmo_id": 1,
    "val_type": "prm_link",
    "constraint": 1,
    "required": False,
    "multiple": True,
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

TPRM_14_FLOAT_DEFAULT_DATA = {
    "name": "Test float TPRM",
    "tmo_id": 1,
    "val_type": "float",
    "required": False,
    "constraint": "0.0:1000.0",
    "created_by": "Test admin",
    "modified_by": "Test admin",
}

INPUT_SHEET_DATA_EXAMPLE = [
    [
        "Test str TPRM",
        "Test int TPRM",
        "Test bool TPRM",
        "Test date TPRM",
        "Test datetime TPRM",
        "Test prm_link TPRM",
        "geometry",
        "Test multiple int TPRM",
        "Test multiple str TPRM",
        "Test multiple float TPRM",
        "Test multiple date TPRM",
        "Test multiple datetime TPRM",
        "Test multiple bool TPRM",
        "Test multiple mo_link TPRM",
        "Test multiple prm_link TPRM",
    ],
    [
        "first_invalid_row",
        "invalid_int_value",
        "1",
        "2021-01-01",
        "2021-01-01 12:30:45",
        "3",
        '{"a": 1}',
        [1, 2, 3, 4, 5],
        ["abc", "def", "ghi"],
        [1.2345],
        ["2021-01-01", "2022-02-02"],
        ["2021-01-01 12:00:00", "2021-01-02 15:30:00"],
        ["true"],
        [],
        None,
    ],
    [
        "second_invalid_string",
        "1.2",
        "WROYEUW",
        "2020-01-04",
        "123",
        "1",
        '{"a": 1}',
        [100],
        ["abc", "def", "ghi"],
        [1.1, 2.2, 3.3],
        ["2023-01-01"],
        None,
        None,
        ["str-1", "str-2", '"parent_object-str-3"'],
        [1, 2],
    ],
    [
        "third_invalid_string",
        "12",
        "1",
        "2020-45-02",
        "2020-12-30 25:61:00",
        "2",
        '{"a": 1}',
        [12, -20, 10],
        ["ABrwrW"],
        [1, 2, 3, 4, 5],
        None,
        ["2022-01-01 00:00:00"],
        ["true", "false", "true", "false"],
        "",
        [],
    ],
    [
        "fourth_invalid_string",
        "4354",
        "1",
        "xbnepgijec",
        "2022-01-01 14:75:00",
        "3",
        '{"a": 1}',
        [123, 20, 20],
        ["A", "B", "C"],
        ["1.1", "2.2", "3.3"],
        ["2021-01-01" for _ in range(10)],
        None,
        ["true", "0"],
        ["str-1", "str-1"],
        "",
    ],
    [
        "fifth_valid_string",
        "193403",
        "true",
        "2022-01-01",
        "2022-01-01 12:30:45",
        "2",
        '{"a": 1}',
        [23434.2, 123.2],
        ["ABC" for _ in range(10)],
        [1.2, 2.3],
        ["2021-01-01", "2022-02-02"],
        None,
        ["1", "1", "1"],
        ["str-1", "str-2"],
        "[]",
    ],
    [
        "sixth_valid_string",
        "193403",
        "true",
        "2022-01-01",
        "2022",
        "ertret",
        '{"a": 1}',
        [True, False, True],
        ["abc", "def", "ghi"],
        [0.1],
        ["2018-08-08"],
        None,
        ["true", "false", "True", "False"],
        None,
        [1, 2],
    ],
    [
        "first_valid_string",
        "5",
        "1",
        "2023-05-10",
        "2023-05-10 10:00:00",
        "14",
        '{"a": 1}',
        [1, 2, 3],
        ["abc"],
        [1.2],
        ["2022-01-01"],
        ["2022-01-01 08:00:00"],
        ["true"],
        ["str-1"],
        [1],
    ],
    [
        "valid_string",
        "5",
        "1",
        "2023-05-10",
        "2023-05-10 10:00:00",
        "14",
        '{"a": 1}',
        [100],
        ["abc", "def", "ghi"],
        [1.1, 2.2, 3.3],
        ["2023-01-01"],
        None,
        None,
        ["str-1", "str-2"],
        [1, 2],
    ],
    [
        "valid_string",
        "5",
        "0",
        "2023-05-10",
        "2023-05-10 10:00:00",
        "14",
        '{"a": 1}',
        [1, 2, 3, 4],
        ["a", "b"],
        [0.1],
        ["2022-01-01"],
        ["2022-01-01 08:15:00"],
        ["false"],
        ["str-5"],
        [5],
    ],
    [
        "valid_string",
        "5",
        "0",
        "2023-05-10",
        "2023-05-10 10:00:00",
        "3",
        '{"a": 1}',
        [3],
        ["def"],
        [0.2],
        ["2023-02-01"],
        ["2023-02-01 12:00:00"],
        ["false"],
        None,
        [2, 3],
    ],
    [
        "second_valid_string",
        "10",
        "0",
        "2002-12-12",
        "2002-12-12 18:45:30",
        "14",
        '{"a": 1}',
        [3, 5],
        ["ghi"],
        [1.3],
        ["24-01-01"],
        ["2024-01-01 14:30:00"],
        ["true"],
        ["parent_object-str-3"],
        [3, 4],
    ],
    [
        "third_valid_string",
        "193403",
        "true",
        "2022-01-01",
        "2022-01-01 08:15:00",
        14,
        '{"a": 1}',
        [100],
        ["xyz"],
        [2.0],
        ["2023-03-01"],
        ["2023-03-01 10:00:00"],
        ["false"],
        ["str-2"],
        [6, 7],
    ],
]


@pytest.fixture(scope="function", autouse=True)
def session_fixture(mocker, session, engine):
    mocker.patch(
        "services.event_service.processor.get_not_auth_session",
        new=lambda: iter([Session(engine)]),
    )
    mocker.patch(
        "services.kafka_service.producer.protobuf_producer.kafka_config.KAFKA_TURN_ON",
        new=False,
    )

    tmo_1 = TMO(**TMO_0_DEFAULT_DATA)
    session.add(tmo_1)
    session.flush()
    tmo = TMO(**TMO_DEFAULT_DATA)
    session.add(tmo)
    session.flush()

    tprm = TPRM(**TPRM_1_STR_DEFAULT_DATA)
    tprm.tmo_id = tmo.id
    session.add(tprm)
    session.flush()

    tprm_1 = TPRM(**TPRM_2_INT_DEFAULT_DATA)
    tprm_1.tmo_id = tmo.id
    session.add(tprm_1)
    session.flush()

    tprm_2 = TPRM(**TPRM_3_BOOL_DEFAULT_DATA)
    tprm_2.tmo_id = tmo.id
    session.add(tprm_2)
    session.flush()

    tprm_3 = TPRM(**TPRM_4_DATE_DEFAULT_DATA)
    tprm_3.tmo_id = tmo.id
    session.add(tprm_3)
    session.flush()

    tprm_4 = TPRM(**TPRM_5_DATETIME_DEFAULT_DATA)
    tprm_4.tmo_id = tmo.id
    session.add(tprm_4)
    session.flush()

    tprm_5 = TPRM(**TPRM_6_MULTI_INT_DEFAULT_DATA)
    tprm_5.tmo_id = tmo.id
    session.add(tprm_5)
    session.flush()

    tprm_6 = TPRM(**TPRM_7_MULTI_STR_DEFAULT_DATA)
    tprm_6.tmo_id = tmo.id
    session.add(tprm_6)
    session.flush()

    tprm_7 = TPRM(**TPRM_8_MULTI_FLOAT_DEFAULT_DATA)
    tprm_7.tmo_id = tmo.id
    session.add(tprm_7)
    session.flush()

    tprm_8 = TPRM(**TPRM_9_MULTI_DATE_DEFAULT_DATA)
    tprm_8.tmo_id = tmo.id
    session.add(tprm_8)
    session.flush()

    tprm_9 = TPRM(**TPRM_10_MULTI_DATETIME_DEFAULT_DATA)
    tprm_9.tmo_id = tmo.id
    session.add(tprm_9)
    session.flush()

    tprm_10 = TPRM(**TPRM_11_MULTI_BOOL_DEFAULT_DATA)
    tprm_10.tmo_id = tmo.id
    session.add(tprm_10)
    session.flush()

    tprm_11 = TPRM(**TPRM_12_MULTI_MO_LINK_DEFAULT_DATA)
    tprm_11.tmo_id = tmo.id
    session.add(tprm_11)
    session.flush()

    tprm_12 = TPRM(**TPRM_13_MULTI_PRM_LINK_DEFAULT_DATA)
    tprm_12.tmo_id = tmo.id
    session.add(tprm_12)
    session.flush()

    tprm_13 = TPRM(**TPRM_14_FLOAT_DEFAULT_DATA)
    tprm_13.tmo_id = tmo.id
    session.add(tprm_13)
    session.flush()

    tprm_14 = TPRM(**TPRM_2_STR_DEFAULT_DATA)
    tprm_14.tmo_id = tmo.id
    session.add(tprm_14)
    session.flush()

    mo_0 = MO(name="parent_object", tmo_id=tmo_1.id, active=True)
    session.add(mo_0)
    session.flush()
    mo_1 = MO(name="str-1", tmo_id=tmo.id, active=True)
    mo_2 = MO(name="str-2", tmo_id=tmo.id, active=True)
    mo_3 = MO(
        name="parent_object-str-3", tmo_id=tmo.id, active=True, p_id=mo_0.id
    )
    mo_4 = MO(name="str-4", tmo_id=tmo_1.id, active=True)
    mo_5 = MO(
        name="parent_object-str-5", tmo_id=tmo.id, active=True, p_id=mo_0.id
    )
    mo_6 = MO(name="parent_object-str-5", tmo_id=tmo_1.id, active=True)
    mo_7 = MO(name="str-6", tmo_id=tmo.id, active=True)

    # session.add(mo_0)
    session.add(mo_1)
    session.add(mo_2)
    session.add(mo_3)
    session.add(mo_4)
    session.add(mo_5)
    session.add(mo_6)
    session.add(mo_7)
    session.flush()

    prm_1 = PRM(mo_id=mo_1.id, tprm_id=tprm.id, value="mo 1 value")
    prm_2 = PRM(mo_id=mo_3.id, tprm_id=tprm.id, value="str")
    prm_3 = PRM(mo_id=mo_3.id, tprm_id=tprm_1.id, value="3")
    prm_4 = PRM(mo_id=mo_3.id, tprm_id=tprm_2.id, value="true")

    prm_5 = PRM(mo_id=mo_7.id, tprm_id=tprm.id, value="str")
    prm_6 = PRM(mo_id=mo_7.id, tprm_id=tprm_1.id, value="6")
    prm_7 = PRM(mo_id=mo_7.id, tprm_id=tprm_3.id, value="2024-01-01")

    session.add(prm_1)
    session.add(prm_2)
    session.add(prm_3)
    session.add(prm_4)

    session.add(prm_5)
    session.add(prm_6)
    session.add(prm_7)

    tprm_5_prm_link = TPRM(
        name="Test prm_link TPRM",
        tmo_id=tmo.id,
        val_type="prm_link",
        constraint=tprm_1.id,
        required=False,
        created_by="Test admin",
        modified_by="Test admin",
    )

    session.add(tprm_5_prm_link)

    session.commit()
    yield session


main_full_data_for_file = [
    ["Test str TPRM", "Test int TPRM", "geometry"],
    ["some_string", "1", '{"a": 1}'],
    ["some_string_2", "2", '{"a": 1}'],
    ["str", "1", '{"a": 1}'],
    ["str", "2", '{"a": 1}'],
]


def test_get_all_needed_sheets(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file = generate_csv_in_memory(main_full_data_for_file)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    assert res.status_code == 200

    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    assert workbook.sheetnames == [
        "Summary",
        "Data Model",
        "Update",
        "New",
        "Errors",
    ]


def test_check_columns_for_summary(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file = generate_csv_in_memory(main_full_data_for_file)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    assert res.status_code == 200

    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)
    summary_sheet = workbook["Summary"]
    headers = [cell.value for cell in summary_sheet[1]]
    assert headers == [
        "Updated objects",
        "Created objects",
        "Updated parameters",
        "Created parameters",
        "Deleted parameters",
    ]


def test_check_columns_for_new(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file = generate_csv_in_memory(main_full_data_for_file)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    assert res.status_code == 200

    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)
    summary_sheet = workbook["New"]
    headers = [cell.value for cell in summary_sheet[1]]
    assert headers == [
        "Parent Name",
        "Object Name",
        "Test str TPRM",
        "Test int TPRM",
        "geometry",
    ]


def test_check_columns_for_update(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file = generate_csv_in_memory(main_full_data_for_file)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    assert res.status_code == 200

    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)
    summary_sheet = workbook["Update"]
    headers = [cell.value for cell in summary_sheet[1]]
    assert headers == [
        "Parent Name",
        "Object Name",
        "Parameter Name",
        "Old Value",
        "New Value",
    ]


def test_check_new_filling_0(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file = generate_csv_in_memory(main_full_data_for_file)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    sheet = workbook["New"]

    # expected headers: ['Object name', 'Parent name', 'Test str TPRM', 'Test int TPRM', 'geometry']
    expected_rows = [
        [None, "some_string-1", "some_string", 1, '{"a": 1}'],
        [None, "some_string_2-2", "some_string_2", 2, '{"a": 1}'],
    ]

    real_rows = [
        [cell.value for cell in sheet[2]],
        [cell.value for cell in sheet[3]],
    ]
    assert expected_rows == real_rows


def test_check_new_filling_with_parents(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file_data_with_parents = [
        ["Test str TPRM", "Test int TPRM", "geometry", "p_id"],
        ["some_string", "1", '{"a": 1}', "1"],
        ["some_string_2", "2", '{"a": 1}', "1"],
        ["str", "1", '{"a": 1}', "1"],
        ["str", "2", '{"a": 1}', "1"],
    ]
    file = generate_csv_in_memory(file_data_with_parents)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    sheet = workbook["New"]

    # expected headers: ['Object name', 'Parent name', 'Test str TPRM', 'Test int TPRM', 'geometry']
    expected_rows = [
        [
            "parent_object",
            "parent_object-some_string-1",
            "some_string",
            1,
            '{"a": 1}',
        ],
        [
            "parent_object",
            "parent_object-some_string_2-2",
            "some_string_2",
            2,
            '{"a": 1}',
        ],
    ]
    real_rows = [
        [cell.value for cell in sheet[2]],
        [cell.value for cell in sheet[3]],
    ]
    assert expected_rows == real_rows


def test_check_update_filling_with_parents(
    session: Session, client: TestClient
):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file_data_with_parents = [
        ["Test str TPRM", "Test int TPRM", "geometry", "p_id"],
        ["str", "5", '{"a": 1}', "1"],
    ]
    file = generate_csv_in_memory(file_data_with_parents)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    sheet = workbook["Update"]

    # expected headers: ['Parent name', 'Object name', 'Parameter Name', 'Old value', 'New Value']
    expected_rows = [
        ["parent_object", "parent_object-str-5", "geometry", None, '{"a": 1}']
    ]
    real_rows = [[cell.value for cell in sheet[2]]]
    print(real_rows)
    assert real_rows == expected_rows


def test_check_update_filling_with_parents_2(
    session: Session, client: TestClient
):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file_data_with_parents = [
        ["Test str TPRM", "Test int TPRM", "geometry", "p_id"],
        ["str", "5", '{"a": 1}', "1"],
    ]
    file = generate_csv_in_memory(file_data_with_parents)
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    sheet = workbook["Update"]

    # expected headers: ['Parent name', 'Object name', 'Parameter Name', 'Old value', 'New Value']
    expected_rows = [
        ["parent_object", "parent_object-str-5", "geometry", None, '{"a": 1}']
    ]
    real_rows = [[cell.value for cell in sheet[2]]]
    print(real_rows)
    assert real_rows == expected_rows


def test_validate_virtual_object_type(session: Session, client: TestClient):
    VIRTUAL_TMO_DATA = {
        "name": "Virtual TMO",
        "version": 1,
        "created_by": "Test Admin",
        "modified_by": "Test Admin",
        "virtual": True,
    }
    virtual_tmo = TMO(**VIRTUAL_TMO_DATA)
    session.add(virtual_tmo)
    session.commit()

    saved_tmo: TMO = session.exec(
        select(TMO).where(TMO.name == "Virtual TMO")
    ).first()

    assert saved_tmo.virtual is True

    file = generate_csv_in_memory(main_full_data_for_file)

    response = client.post(
        url=(URL + f"{saved_tmo.id}"),
        data={
            "filename": file.name,
            "type": "multipart/form-data",
        },
        files={"file": file},
    )

    assert response.status_code == 400
    assert (
        response.json()["detail"]
        == f"TMO ({virtual_tmo.name}) is virtual. You can`t create objects with TMO.virtual equal to True."
    )


def test_validate_invalid_file_content_type(
    session: Session, client: TestClient
):
    valid_tmo = session.exec(select(TMO).where(TMO.id == 2)).first()

    invalid_file_content = b"Invalid content for testing"
    invalid_file_name = "invalid_file.txt"
    invalid_file = BytesIO(invalid_file_content)

    response = client.post(
        url=(URL + f"{valid_tmo.id}"),
        data={
            "filename": invalid_file_name,
            "type": "multipart/form-data",
        },
        files={
            "file": invalid_file,
        },
    )
    print(response.json())
    assert response.status_code == 422
    assert "Allowed content types" in response.json()["detail"]


def test_validate_non_unique_file_headers(session: Session, client: TestClient):
    res = session.exec(select(TMO).where(TMO.id == 2)).first()

    file_data_with_non_unique_headers = [
        ["Test str TPRM", "Test str TPRM", "Test int TPRM"],  # Repeated headers
        ["some_string", "some_string_2", "1"],
        ["some_string_3", "some_string_4", "2"],
    ]

    file = generate_csv_in_memory(file_data_with_non_unique_headers)

    response = client.post(
        URL + f"{res.id}",
        data={
            "filename": file.name,
            "type": "multipart/form-data",
        },
        files={"file": file},
    )
    assert response.status_code == 422
    expected_error_message = (
        "Column names must be unique. Not unique columns: ['Test str TPRM']."
    )
    assert response.json()["detail"] == expected_error_message


def test_validate_multiple_groups_of_non_unique_headers(
    session: Session, client: TestClient
):
    tmo_instance = session.exec(select(TMO).where(TMO.id == 2)).first()

    file_data_with_non_unique_headers = [
        [
            "Test str TPRM",
            "Test int TPRM",
            "Test str TPRM",
            "geometry",
            "location.x",
            "location.y",
            "geometry",
        ],
        [
            "some_string",
            "1",
            "some_string_2",
            '{"a": 1}',
            "100",
            "200",
            '{"b": 2}',
        ],
        [
            "some_string_3",
            "2",
            "some_string_4",
            '{"b": 2}',
            "101",
            "201",
            '{"c": 3}',
        ],
    ]

    file = generate_csv_in_memory(file_data_with_non_unique_headers)

    response = client.post(
        URL + f"{tmo_instance.id}",
        data={
            "filename": file.name,
            "type": "multipart/form-data",
        },
        files={
            "file": file,
        },
    )
    expected_error_message_part = (
        "Column names must be unique. Not unique columns:"
    )
    assert response.status_code == 422
    assert expected_error_message_part in response.json()["detail"]


def test_update_header_with_nonexistent_tprm(
    session: Session, client: TestClient
):
    tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()

    file_data = [
        [
            "nonexistent_name",
            "Test int TPRM",
            "geometry",
        ],  # 'nonexistent_name' does not exist
        ["some_string", "1", '{"a": 1}'],
        ["some_string_2", "2", '{"a": 1}'],
    ]

    file = generate_csv_in_memory(file_data)

    column_name_mapping = {
        "nonexistent_name": "Invalid TPRM Name"  # Invalid name
    }

    response = client.post(
        URL + f"{tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
        json=column_name_mapping,
    )

    assert response.status_code == 422
    assert "not exists" in response.json()["detail"]


def test_update_header_with_mixed_and_mo_attributes(
    session: Session, client: TestClient
):
    tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()

    # Create and add TPRM records for the test
    tprm_str = TPRM(
        name="Test str TPRM 2",
        val_type="str",
        required=True,
        created_by="Test admin",
        modified_by="Test admin",
        tmo_id=tmo.id,  # Associate with TMO
    )

    tprm_int = TPRM(
        name="Test int TPRM 2",
        val_type="int",
        required=True,
        created_by="Test admin",
        modified_by="Test admin",
        tmo_id=tmo.id,  # Associate with TMO
    )

    session.add(tprm_str)
    session.add(tprm_int)
    session.commit()

    assert (
        session.exec(select(TPRM).where(TPRM.name == "Test str TPRM 2")).first()
        is not None
    )
    assert (
        session.exec(select(TPRM).where(TPRM.name == "Test int TPRM 2")).first()
        is not None
    )

    file_data = [
        [
            "valid_name_1",
            "nonexistent_name",
            "Test int TPRM 2",
            "geometry",
            "active",
            "parent_name",
        ],
        ["some_string", "1", "2", '{"a": 1}', "True", "Parent1"],
        ["some_string_2", "3", "4", '{"b": 2}', "False", "Parent2"],
    ]

    file = generate_csv_in_memory(file_data)

    column_name_mapping = {
        "valid_name_1": "Test str TPRM 2",  # Valid mapping
        "Test int TPRM 2": "Test int TPRM 2",  # Valid and unchanged
        "nonexistent_name": "Invalid TPRM Name",  # Invalid mapping (mapped name does not exist)
        "another_invalid_name": "Invalid Another",  # Invalid mapping, but not in headers
    }

    response = client.post(
        URL + f"{tmo.id}",
        data={
            "filename": file.name,
            "type": "multipart/form-data",
            "column_name_mapping": json.dumps(column_name_mapping),
        },
        files={"file": file},
    )

    assert response.status_code == 422

    error_detail = response.json()["detail"]
    print(error_detail)

    assert "not exists" in error_detail
    assert "Invalid TPRM Name" in error_detail

    for mo_attr in ["geometry", "status", "active", "parent_name"]:
        assert mo_attr not in error_detail


def test_batch_test_NA_values_in_data(session: Session, client: TestClient):
    tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()

    file_data_with_nan = [
        ["Test str TPRM", "Test int TPRM", "Test str_1 TPRM"],
        ["some_string_1", "1", "some_value"],
        ["some_string_2", "2", "N/A"],
        ["some_string_3", "3", None],
        ["some_string_4", "4", "N/A"],
    ]

    file = generate_csv_in_memory(file_data_with_nan)

    res = client.post(
        URL + f"{tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    assert res.status_code == 200

    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    sheet = workbook["New"]
    actual_values = []
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, values_only=True
    ):
        actual_values.append([row])
    expected_values = [
        [
            (
                "Parent Name",
                "Object Name",
                "Test str TPRM",
                "Test int TPRM",
                "Test str_1 TPRM",
            )
        ],
        [(None, "some_string_1-1", "some_string_1", 1, "some_value")],
        [(None, "some_string_2-2", "some_string_2", 2, "N/A")],
        [(None, "some_string_3-3", "some_string_3", 3, None)],
        [(None, "some_string_4-4", "some_string_4", 4, "N/A")],
    ]

    assert expected_values == actual_values


def test_batch_objects_preview_missing_required_fields(
    client: TestClient, session: Session
):
    tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()

    # Create and add TPRM records for the test
    tprm_str = TPRM(
        name="Test str TPRM 2",
        val_type="str",
        required=True,
        created_by="Test admin",
        modified_by="Test admin",
        tmo_id=tmo.id,  # Associate with TMO
    )

    tprm_int = TPRM(
        name="Test int TPRM 2",
        val_type="int",
        required=True,
        created_by="Test admin",
        modified_by="Test admin",
        tmo_id=tmo.id,  # Associate with TMO
    )

    # Add TPRMs to the session and commit
    session.add(tprm_str)
    session.add(tprm_int)
    session.commit()

    assert (
        session.exec(select(TPRM).where(TPRM.name == "Test str TPRM 2")).first()
        is not None
    )
    assert (
        session.exec(select(TPRM).where(TPRM.name == "Test int TPRM 2")).first()
        is not None
    )

    file_data = [
        [
            "Test str TPRM",
            "Test int TPRM 2",
            "geometry",
            "active",
            "parent_name",
            "asd",
        ],
        ["some_string", "2", '{"a": 1}', "True", "Parent1", "qewe"],
        ["some_string_2", "4", '{"b": 2}', "False", "Parent2", "qew"],
    ]

    file = generate_csv_in_memory(file_data)

    # Send POST request to the API
    response = client.post(
        URL + f"{tmo.id}",
        data={
            "filename": file.name,
            "type": "multipart/form-data",
        },
        files={"file": file},
    )

    assert response.status_code == 422

    error_detail = response.json()["detail"]
    print(error_detail)

    assert "not exists" in error_detail

    for mo_attr in ["geometry", "status", "active", "parent_name"]:
        assert mo_attr not in error_detail


### 7. Then we dive into `self.__validate_file_values()` which is located next to validate_headers_for_file()
###    call. Inside this method we. This method consists of the following actions:
###       1) validate mo_attr or tprm values column by column, and get errors/warnings;
###       2) update self.__main_dataframe respectively;
###       3) validate sequence tprm separately in the method called `self.__sequence_validation()`
###    Test cases:
###       1) several invalid TPRMs (different value type);
###       2) several invalid attributes (different types);
###       3) several invalid attrs and prms (everything is different);
###       4) everything is works correctly.


# def test_validate_invalid_single_tprm_values(
#     session: Session, client: TestClient
# ):
#     """The test mainly devoted to check for single TPRMs. TODO: Add some new types checking"""
#     tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()
#
#     file_data_with_invalid_tprm_values = [
#         [
#             "Test str TPRM",
#             "Test int TPRM",
#             "Test bool TPRM",
#             "Test date TPRM",
#             "Test datetime TPRM",
#             "Test prm_link TPRM",
#             "Test float TPRM",
#         ],
#         [
#             "first_invalid_row",
#             "invalid_int_value",
#             "1",
#             "2021-01-01",
#             "2021-01-01 12:30:45",
#             "3",
#             None,
#         ],
#         [
#             "second_invalid_string",
#             "1.2",
#             "WROYEUW",
#             "2020-01-04",
#             "123",
#             "1",
#             "true",
#         ],
#         [
#             "third_invalid_string",
#             "12",
#             "1",
#             "2020-45-02",
#             "2020-12-30 25:61:00",
#             "2",
#             "0.0.0.0",  # noqa: S104
#         ],
#         [
#             "fourth_invalid_string",
#             "4354",
#             "1",
#             "xbnepgijec",
#             "2022-01-01 14:75:00",
#             "3",
#             "!@#$",
#         ],
#         [
#             "fifth_valid_string",
#             "193403",
#             "true",
#             "2022-01-01",
#             "2022-01-01 12:30:45",
#             "2",
#             None,
#         ],
#         [
#             "sixth_valid_string",
#             "193403",
#             "true",
#             "2022-01-01",
#             "2022",
#             "ertret",
#             "aezakmi",
#         ],
#         [
#             "valid_string",
#             "193.0.3",
#             "true",
#             "2022-01-01",
#             "2022-01-01 12:30:45",
#             " 2 ",
#             "123.12.12",
#         ],
#         # Valid rows
#         [
#             "first_valid_string",
#             "5",
#             "true",
#             "2023-05-10",
#             None,
#             "15",
#             "2040.12234",
#         ],
#         [
#             "valid_string",
#             "5",
#             None,
#             "2023-05-10",
#             "2023-05-10 10:00:00",
#             "15",
#             "1",
#         ],
#         [
#             "valid_string",
#             "5",
#             "true",
#             "2023-05-10",
#             "2023-05-10 10:00:00",
#             "15",
#             1.1,
#         ],
#         [
#             "valid_string",
#             "5",
#             "true",
#             None,
#             "2023-05-10 10:00:00",
#             "15",
#             "1.01",
#         ],
#         [
#             "second_valid_string",
#             "10",
#             "0",
#             "2002-12-12",
#             "2002-12-12 18:45:30",
#             None,
#             "12.12",
#         ],
#         [
#             "third_valid_string",
#             "193403",
#             "true",
#             "2022-01-01",
#             "2022-01-01 08:15:00",
#             "15",
#             "0.0",
#         ],
#     ]
#
#     file = generate_csv_in_memory(file_data_with_invalid_tprm_values)
#
#     response = client.post(
#         URL + f"{tmo.id}",
#         data={"filename": file.name, "type": "multipart/form-data"},
#         files={"file": file},
#     )
#     assert response.status_code == 200
#     response_stream = BytesIO(response.content)
#     response_stream.seek(0)
#
#     workbook = openpyxl.load_workbook(response_stream)
#     # workbook.save('workbook.xlsx')
#     error_sheet = workbook["Errors"]
#
#     expected_errors = [
#         [
#             "invalid_int_value",
#             "Test int TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "193.0.3",
#             "Test int TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "WROYEUW",
#             "Test bool TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "2020-45-02",
#             "Test date TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "xbnepgijec",
#             "Test date TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "123",
#             "Test datetime TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "2020-12-30 25:61:00",
#             "Test datetime TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "2022-01-01 14:75:00",
#             "Test datetime TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "2022",
#             "Test datetime TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "ertret",
#             "Test prm_link TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             " 2 ",
#             "Test prm_link TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "valid_string-5",
#             "Object Name",
#             "This object name appears more than once at indexes: [8, 9, 10]",
#         ],
#         [
#             "true",
#             "Test float TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "123.12.12",
#             "Test float TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "0.0.0.0",  # noqa: S104
#             "Test float TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "aezakmi",
#             "Test float TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "!@#$",
#             "Test float TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         ["1", "Test prm_link TPRM", "You try to add not exists object"],
#         ["2", "Test prm_link TPRM", "You try to add not exists object"],
#         ["15", "Test prm_link TPRM", "You try to add not exists object"],
#         [
#             "0.0",
#             "Test float TPRM",
#             "Current TPRM has constraint, but this values doesn't match it",
#         ],
#         [
#             "2040.12234",
#             "Test float TPRM",
#             "Current TPRM has constraint, but this values doesn't match it",
#         ],
#     ]
#
#     actual_errors = list()
#
#     for row in error_sheet.iter_rows(
#         min_row=2, max_row=error_sheet.max_row, values_only=True
#     ):
#         value, tprm, error_message = row[3], row[2], row[4]
#         if value and tprm and error_message:
#             actual_errors.append([value, tprm, error_message])
#
#     # print(actual_errors)
#     actual_errors_set = set(tuple(e) for e in actual_errors)
#     expected_errors_set = set(tuple(e) for e in expected_errors)
#
#     extra_in_actual = actual_errors_set - expected_errors_set
#     missing_in_actual = expected_errors_set - actual_errors_set
#
#     print(f"Extra errors in actual: {extra_in_actual}")
#     print(f"Missing errors in actual: {missing_in_actual}")
#     assert not extra_in_actual and not missing_in_actual, (
#         f"Found discrepancies. Extra in actual: {extra_in_actual}. Missing in actual: {missing_in_actual}"
#     )
#

# def test_validate_invalid_multi_tprm_values(
#     session: Session, client: TestClient
# ):
#     """The test is designed to test the creation of multiple TPRMs"""
#     tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()
#     file_data_with_invalid_tprm_values = [
#         [
#             "Test multiple int TPRM",
#             "Test multiple str TPRM",
#             "Test multiple float TPRM",
#             "Test multiple date TPRM",
#             "Test multiple datetime TPRM",
#             "Test multiple bool TPRM",
#             "Test multiple mo_link TPRM",
#             "Test multiple prm_link TPRM",
#             "Test str TPRM",
#             "Test int TPRM",
#         ],
#         # Valid
#         [
#             [1, 2, 3, 4, 5],
#             ["abc", "def", "ghi"],
#             [1.2345],
#             ["2021-01-01", "2022-02-02"],
#             ["2021-01-01 12:00:00", "2021-01-02 15:30:00"],
#             ["true"],
#             [],
#             None,
#             "str",
#             "1",
#         ],
#         [
#             [100],
#             "['abc', 'def', 'ghi']",
#             [1.1, 2.2, 3.3],
#             ["2023-01-01"],
#             None,
#             None,
#             ["str-1", "str-2", "parent_object-str-3"],
#             [1, 2],
#             "str_2",
#             "1",
#         ],
#         [
#             [12, -20, 10],
#             ["ABrwrW"],
#             [1, 2, 3, 4, 5],
#             None,
#             ["2022-01-01 00:00:00"],
#             ["true", "false", "true", "false"],
#             "",
#             [1],
#             "str_3",
#             "1",
#         ],
#         [
#             "[123, 20, 20]",
#             ["A", "B", "C"],
#             ["1.1", "2.2", "3.3"],
#             ["2021-01-01" for _ in range(10)],
#             None,
#             ["true", "0"],
#             ["str-1", "str-1"],
#             "",
#             "str_4",
#             "1",
#         ],
#         [
#             [23434.2, 123.2],
#             ["ABC" for _ in range(10)],
#             [1.2, 2.3],
#             ["2021-01-01", "2022-02-02"],
#             None,
#             ["1", "1", "1"],
#             ["str-1", "str-2"],
#             "[]",
#             "str_5",
#             "1",
#         ],
#         [
#             [True, False, True],
#             ["abc", "def", "ghi"],
#             [0.1],
#             ["2018-08-08"],
#             None,
#             '["true", "false", "True", "False"]',
#             None,
#             "[1, 2]",
#             "str_6",
#             "1",
#         ],
#         # Invalid
#         [
#             ["true", 10],
#             ["ab!c", 1, "ghi"],
#             [-1.1, -2.2],
#             ["2021-01-01", "2022-02-33"],
#             ["2024-01-01 32:00:00"],
#             None,
#             None,
#             [],
#             "str_7",
#             "1",
#         ],
#         [
#             ["true", "12323.!!"],
#             True,
#             "aezakmi",
#             [True],
#             [True, False, True, 1],
#             None,
#             ["str-15"],
#             [1, 2, 3],
#             "str_8",
#             "1",
#         ],
#         [
#             [1, 2],
#             ["a", "b"],
#             ["aezakmi", "werwer"],
#             "234324234",
#             None,
#             ("TrUe", "1", "2", True),
#             True,
#             "whworho",
#             "str_9",
#             "1",
#         ],
#         [
#             [100],
#             "['abc', 'def', 'ghi']",
#             [1.1, 2.2, 3.3],
#             ["12000-01-01"],
#             None,
#             None,
#             ["str-6", "str-1", None],
#             [1, 10.2, 2.102],
#             "str_10",
#             "1",
#         ],
#         [
#             [100],
#             ["abc", "def", "ghi"],
#             [1.1, 2.2, 3.3],
#             ["2022-02-25 14:20:00"],
#             ["2024-01-01"],
#             '"werer"',
#             None,
#             [True, False, 0],
#             "str_11",
#             "1",
#         ],
#         [
#             [100],
#             ["abc", "def", "ghi"],
#             [1.1, 2.2, 3.3],
#             ["2022-02-25 14:20:00"],
#             ["2024-01-01"],
#             [1, 1, 1],
#             ["str-4", "str-1"],
#             [1, 2, 3, 4],
#             "str_12",
#             "1",
#         ],
#         [
#             [100],
#             "['abc', 'def', 'ghi']",
#             [1.1, 2.2, 3.3],
#             ["12000-01-01"],
#             None,
#             ["qwerty"],
#             [None, "str-1", None],
#             ["23432", "wrihew"],
#             "str_13",
#             "1",
#         ],
#         [
#             [100],
#             ["abc", "def"],
#             [2.2, 3.3],
#             ["2022-02-25 14:20:00"],
#             ["2024-01-01"],
#             [1],
#             ["parent_object-str-5"],
#             None,
#             "str_14",
#             "1",
#         ],
#     ]
#
#     file = generate_csv_in_memory(file_data_with_invalid_tprm_values)
#
#     response = client.post(
#         URL + f"{tmo.id}",
#         data={"filename": file.name, "type": "multipart/form-data"},
#         files={"file": file},
#     )
#     assert response.status_code == 200
#     response_stream = BytesIO(response.content)
#     response_stream.seek(0)
#
#     workbook = openpyxl.load_workbook(response_stream)
#     # workbook.save('workbook.xlsx')
#     error_sheet = workbook["Errors"]
#
#     actual_errors = list()
#
#     for row in error_sheet.iter_rows(
#         min_row=2, max_row=error_sheet.max_row, values_only=True
#     ):
#         value, tprm, error_message = row[3], row[2], row[4]
#         if value and tprm and error_message:
#             actual_errors.append([value, tprm, error_message])
#
#     # print(actual_errors)
#     expected_errors = [
#         [
#             "['true', '12323.!!']",
#             "Test multiple int TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "['ab!c', 1, 'ghi']",
#             "Test multiple str TPRM",
#             "Current TPRM has constraint, but this values doesn't match it",
#         ],
#         [
#             "True",
#             "Test multiple str TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         ### Another report message
#         [
#             "aezakmi",
#             "Test multiple float TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         [
#             "['aezakmi', 'werwer']",
#             "Test multiple float TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "['2021-01-01', '2022-02-33']",
#             "Test multiple date TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],  ### There should be another message here
#         [
#             "['12000-01-01']",
#             "Test multiple date TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],  ### as well
#         [
#             "234324234",
#             "Test multiple date TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         ### as well
#         [
#             "[True]",
#             "Test multiple date TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         ### as well
#         [
#             "['2022-02-25 14:20:00']",
#             "Test multiple date TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],  ### as well
#         [
#             "[True, False, True, 1]",
#             "Test multiple datetime TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],  ### as well
#         [
#             "['2024-01-01 32:00:00']",
#             "Test multiple datetime TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],  ### as well
#         [
#             "['2024-01-01']",
#             "Test multiple datetime TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],  ### as well
#         [
#             "['qwerty']",
#             "Test multiple bool TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             '"werer"',
#             "Test multiple bool TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         [
#             "('TrUe', '1', '2', True)",
#             "Test multiple bool TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         [
#             "['str-15']",
#             "Test multiple mo_link TPRM",
#             "You try to add not exists object",
#         ],
#         [
#             "['str-6', 'str-1', None]",
#             "Test multiple mo_link TPRM",
#             "You try to add not exists object",
#         ],
#         [
#             "[None, 'str-1', None]",
#             "Test multiple mo_link TPRM",
#             "You try to add not exists object",
#         ],
#         [
#             "True",
#             "Test multiple mo_link TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         [
#             "['parent_object-str-5']",
#             "Test multiple mo_link TPRM",
#             "Object name is ambiguous! Provide more concrete name!",
#         ],
#         [
#             "[1, 2, 3]",
#             "Test multiple prm_link TPRM",
#             "You try to add not exists object",
#         ],
#         [
#             "whworho",
#             "Test multiple prm_link TPRM",
#             "Current TPRM is multiple, so it has to store list of values",
#         ],
#         [
#             "['23432', 'wrihew']",
#             "Test multiple prm_link TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "[1, 2, 3, 4]",
#             "Test multiple prm_link TPRM",
#             "You try to add not exists object",
#         ],
#         # ['[1, 2]', 'Test multiple prm_link TPRM', 'You try to add not exists object'],
#         [
#             "[True, False, 0]",
#             "Test multiple prm_link TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         [
#             "[1, 10.2, 2.102]",
#             "Test multiple prm_link TPRM",
#             "Every TPRM has own value type, but this value doesn't match it",
#         ],
#         # ["['1', '2']", 'Test multiple prm_link TPRM', 'You try to add not exists object'],
#         # ["['GOAT']", 'Test multiple prm_link TPRM', 'Current TPRM is multiple, so it has to store list of values'],
#     ]
#
#     actual_errors_set = set(tuple(e) for e in actual_errors)
#     expected_errors_set = set(tuple(e) for e in expected_errors)
#
#     extra_in_actual = actual_errors_set - expected_errors_set
#     missing_in_actual = expected_errors_set - actual_errors_set
#
#     print(f"Extra errors in actual: {extra_in_actual}")
#     print(f"Missing errors in actual: {missing_in_actual}")
#
#     assert not extra_in_actual and not missing_in_actual, (
#         f"Found discrepancies. Extra in actual: {extra_in_actual}. Missing in actual: {missing_in_actual}"
#     )


def test_validate_invalid_attribute_values(
    session: Session, client: TestClient
):
    """
    The test is designed to test the creation of attribute values.
    TODO: check 'point_b_name' validation.
    """
    tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()

    file_data_with_invalid_attribute_values = [
        [
            "geometry",
            "active",
            "pov",
            "p_id",
            "parent_name",
            "point_a_name",
            "Test str TPRM",
            "Test int TPRM",
        ],
        [None, "true", '{"a": 1}', "123", "invalid_parent", "", "str", "1"],
        ['{"b": 1}', "false", '{"pov": 2}', "", None, None, "str_1", "1"],
        # ['', ''], # unexpectable error
        [
            '{"d": 1}',
            "12#",
            "invalid pov",
            "p-123",
            "unknown_parent",
            None,
            "str_2",
            "1",
        ],
        [
            '{"e": 1}',
            "1",
            '{"b": null}',
            "invalid_id!",
            "invalid_parent2",
            None,
            "str_3",
            "1",
        ],
        [
            '{"c": null}',
            "falsew",
            '{"invalid": 1}',
            "p_id%",
            "",
            "",
            "str_4",
            "1",
        ],
        ['{"f": 1}', "1", '{"pov": "xyz"}', "1", 123, "1", "str_5", "1"],
        [
            '{"f": 1, "a": 234}',
            "1",
            '{"pov": 123}',
            123,
            "another_invalid_parent",
            None,
            "str_6",
            "1",
        ],
        [
            '{"f": 1, "abc": "234"}',
            1,
            '{"x": "test"}',
            1,
            "parent_object",
            None,
            "str_7",
            "1",
        ],
        [
            '{"f": 1, "abc": 234,}',
            "false",
            "pov data",
            " ",
            "",
            None,
            "str_12",
            "1",
        ],
        [
            "[true, false]",
            "true",
            '{"pov": [1, 2]}',
            "abc123",
            None,
            None,
            "str_8",
            "1",
        ],
        [
            "f 1",
            False,
            '{"pov": null}',
            None,
            "parent_name_with_special_char@",
            None,
            "str_9",
            "1",
        ],
        [
            "12.2313123/45345",
            "1",
            '{"a": "invalid"}',
            "!!!",
            "parent1",
            None,
            "str_132",
            "1",
        ],
        ["12.2313123/45345", "1", "pov1", "0000", "", None, "str_10", "1"],
        [
            '{"d": 1, "u": 2, "c": 3, "k": 4}',
            "1",
            "invalid pov",
            "p id",
            "invalid_parent_name",
            "",
            "str_11",
            "1",
        ],
    ]

    file = generate_csv_in_memory(file_data_with_invalid_attribute_values)

    response = client.post(
        URL + f"{tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    assert response.status_code == 200
    response_stream = BytesIO(response.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)
    # workbook.save('error_sheet.xlsx')
    error_sheet = workbook["Errors"]

    expected_errors = [
        [
            " ",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            '{"f": 1, "abc": 234,}',
            "geometry",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "[true, false]",
            "geometry",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "f 1",
            "geometry",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "12.2313123/45345",
            "geometry",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "12#",
            "active",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "falsew",
            "active",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "invalid pov",
            "pov",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "pov data",
            "pov",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "pov1",
            "pov",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "p-123",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "invalid_id!",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "p_id%",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "abc123",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "!!!",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        [
            "p id",
            "p_id",
            "Every attribute has own value type, but this value doesn't match it",
        ],
        ["0000", "p_id", "You try to add not exists object"],
        ["123", "p_id", "You try to add not exists object"],
        ["unknown_parent", "parent_name", "You try to add not exists object"],
        ["parent1", "parent_name", "You try to add not exists object"],
        ["123", "parent_name", "You try to add not exists object"],
        [
            "another_invalid_parent",
            "parent_name",
            "You try to add not exists object",
        ],
        [
            "invalid_parent_name",
            "parent_name",
            "You try to add not exists object",
        ],
        [
            "parent_name_with_special_char@",
            "parent_name",
            "You try to add not exists object",
        ],
        ["invalid_parent", "parent_name", "You try to add not exists object"],
        ["invalid_parent2", "parent_name", "You try to add not exists object"],
        [
            "1",
            "point_a_name",
            "Current TPRM has constraint, but this values doesn't match it",
        ],
    ]

    actual_errors = list()

    for row in error_sheet.iter_rows(
        min_row=2, max_row=error_sheet.max_row, values_only=True
    ):
        value, tprm, error_message = row[3], row[2], row[4]
        if value and tprm and error_message:
            actual_errors.append([value, tprm, error_message])

    # print(actual_errors)
    actual_errors_set = set(tuple(e) for e in actual_errors)
    expected_errors_set = set(tuple(e) for e in expected_errors)

    extra_in_actual = actual_errors_set - expected_errors_set
    missing_in_actual = expected_errors_set - actual_errors_set

    print(f"Extra errors in actual: {extra_in_actual}")
    print(f"Missing errors in actual: {missing_in_actual}")

    assert not extra_in_actual and not missing_in_actual, (
        f"Found discrepancies. Extra in actual: {extra_in_actual}. Missing in actual: {missing_in_actual}"
    )


def test_update_mos_reports(session: Session, client: TestClient):
    """The test is about reports creating when update MOs."""
    tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()

    file_data_with_invalid_tprm_values = [
        [
            "Test str TPRM",
            "Test int TPRM",
            "Test bool TPRM",
            "Test date TPRM",
            "Test datetime TPRM",
            "parent_name",
        ],
        [
            "str",
            "3",
            "false",
            "2021-01-01",
            "202-01- 12:30:45",
            "parent_object",
        ],
        [
            "str",
            "3",
            "incorrect_bool",
            "2021-01-01",
            "2021-01-01 12:30:45",
            "parent_object",
        ],
        ["str", "6", None, None, None, None],
        ["str", "6", None, None, None, None],
        ["str", "6", None, None, None, None],
        ["new_str", "11", "aezakmi", "2021-01-01", "2021-01-01 12:30:45", None],
        [
            "new_str",
            "13",
            "incorrect_bool",
            "2021-01-01",
            "2021-01-01 12:30:45",
            None,
        ],
    ]

    file = generate_csv_in_memory(file_data_with_invalid_tprm_values)

    response = client.post(
        URL + f"{tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )

    assert response.status_code == 200
    response_stream = BytesIO(response.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)
    # workbook.save('workbook.xlsx')

    actual_summary_headers_and_content = list()
    summary_sheet = workbook["Summary"]
    max_row = summary_sheet.max_row

    for row in summary_sheet.iter_rows(
        min_row=1, max_row=max_row, values_only=True
    ):
        actual_summary_headers_and_content.append(list(row))

    # print(actual_summary_headers_and_content)
    expected_summary_headers_and_content = [
        [
            "Updated objects",
            "Created objects",
            "Updated parameters",
            "Created parameters",
            "Deleted parameters",
        ],
        [2, 2, 1, 9, 1],
    ]

    assert (
        actual_summary_headers_and_content
        == expected_summary_headers_and_content
    )

    actual_update_content = list()
    update_sheet = workbook["Update"]
    max_row = update_sheet.max_row

    for row in update_sheet.iter_rows(
        min_row=2,
        max_row=max_row,
        values_only=True,
    ):
        parent_name, object_name, prm_name, old_val, new_val = row
        actual_update_content.append(
            [
                parent_name,
                object_name,
                prm_name,
                old_val,
                new_val,
            ]
        )

    expected_update_sheet_content = [
        [
            "parent_object",
            "parent_object-str-3",
            "Test bool TPRM",
            "true",
            False,
        ],
        [
            "parent_object",
            "parent_object-str-3",
            "Test date TPRM",
            None,
            "2021-01-01",
        ],
    ]

    assert actual_update_content == expected_update_sheet_content

    actual_errors_content = list()
    errors_sheet = workbook["Errors"]
    max_row = errors_sheet.max_row

    for row in errors_sheet.iter_rows(
        min_row=2,
        max_row=max_row,
        values_only=True,
    ):
        parent_name, object_name, prm_name, value, reason, row_id = row
        actual_errors_content.append(
            [parent_name, object_name, prm_name, value, reason, row_id]
        )

    expected_errors_sheet_content = [
        [
            None,
            "new_str-11",
            "Test bool TPRM",
            "aezakmi",
            "Every TPRM has own value type, but this value doesn't match it",
            6,
        ],
        [
            None,
            "new_str-13",
            "Test bool TPRM",
            "incorrect_bool",
            "Every TPRM has own value type, but this value doesn't match it",
            7,
        ],
        [
            "parent_object",
            "parent_object-str-3",
            "Test datetime TPRM",
            "202-01- 12:30:45",
            "Every TPRM has own value type, but this value doesn't match it",
            1,
        ],
        [
            "parent_object",
            "parent_object-str-3",
            "Object Name",
            "parent_object-str-3",
            "This object name appears more than once at indexes: [0, 1]",
            1,
        ],
        [
            None,
            "str-6",
            "Object Name",
            "str-6",
            "This object name appears more than once at indexes: [2, 3, 4]",
            3,
        ],
        [
            None,
            None,
            "Test bool TPRM",
            "incorrect_bool",
            "Every TPRM has own value type, but this value doesn't match it",
            2,
        ],
    ]

    assert actual_errors_content == expected_errors_sheet_content

    # def test_validate_data_model_sheet_types(session: Session, client: TestClient):
    #     """The test mainly devoted to data model sheet creation."""
    #     tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()
    #     tprm_prm_link = session.exec(
    #         select(TPRM).where(TPRM.name == "Test prm_link TPRM")
    #     ).first()
    #
    #     file_data = [
    #         [
    #             "Test str TPRM",
    #             "Test int TPRM",
    #             "Test bool TPRM",
    #             "Test date TPRM",
    #             "Test datetime TPRM",
    #             "Test prm_link TPRM",
    #             "geometry",
    #             "Test multiple int TPRM",
    #             "Test multiple str TPRM",
    #             "Test multiple float TPRM",
    #             "Test multiple date TPRM",
    #             "Test multiple datetime TPRM",
    #             "Test multiple bool TPRM",
    #             "Test multiple mo_link TPRM",
    #             "Test multiple prm_link TPRM",
    #         ],
    #         [
    #             "first_invalid_row",
    #             "invalid_int_value",
    #             "1",
    #             "2021-01-01",
    #             "2021-01-01 12:30:45",
    #             "3",
    #             '{"a": 1}',
    #             [1, 2, 3, 4, 5],
    #             ["abc", "def", "ghi"],
    #             [1.2345],
    #             ["2021-01-01", "2022-02-02"],
    #             ["2021-01-01 12:00:00", "2021-01-02 15:30:00"],
    #             ["true"],
    #             [],
    #             None,
    #         ],
    #         [
    #             "second_invalid_string",
    #             "1.2",
    #             "WROYEUW",
    #             "2020-01-04",
    #             "123",
    #             "1",
    #             '{"a": 1}',
    #             [100],
    #             ["abc", "def", "ghi"],
    #             [1.1, 2.2, 3.3],
    #             ["2023-01-01"],
    #             None,
    #             None,
    #             ["str-1", "str-2", "str-3"],
    #             [1, 2],
    #         ],
    #         [
    #             "third_invalid_string",
    #             "12",
    #             "1",
    #             "2020-45-02",
    #             "2020-12-30 25:61:00",
    #             "2",
    #             '{"a": 1}',
    #             [12, -20, 10],
    #             ["ABrwrW"],
    #             [1, 2, 3, 4, 5],
    #             None,
    #             ["2022-01-01 00:00:00"],
    #             ["true", "false", "true", "false"],
    #             "",
    #             [],
    #         ],
    #         [
    #             "fourth_invalid_string",
    #             "4354",
    #             "1",
    #             "xbnepgijec",
    #             "2022-01-01 14:75:00",
    #             "3",
    #             '{"a": 1}',
    #             [123, 20, 20],
    #             ["A", "B", "C"],
    #             ["1.1", "2.2", "3.3"],
    #             ["2021-01-01" for _ in range(10)],
    #             None,
    #             ["true", "0"],
    #             ["str-1", "str-1"],
    #             "",
    #         ],
    #         [
    #             "fifth_valid_string",
    #             "193403",
    #             "true",
    #             "2022-01-01",
    #             "2022-01-01 12:30:45",
    #             "2",
    #             '{"a": 1}',
    #             [23434.2, 123.2],
    #             ["ABC" for _ in range(10)],
    #             [1.2, 2.3],
    #             ["2021-01-01", "2022-02-02"],
    #             None,
    #             ["1", "1", "1"],
    #             ["str-1", "str-2"],
    #             "[]",
    #         ],
    #         [
    #             "sixth_valid_string",
    #             "193403",
    #             "true",
    #             "2022-01-01",
    #             "2022",
    #             "ertret",
    #             '{"a": 1}',
    #             [True, False, True],
    #             ["abc", "def", "ghi"],
    #             [0.1],
    #             ["2018-08-08"],
    #             None,
    #             ["true", "false", "True", "False"],
    #             None,
    #             [1, 2],
    #         ],
    #         [
    #             "first_valid_string",
    #             "5",
    #             "1",
    #             "2023-05-10",
    #             "2023-05-10 10:00:00",
    #             str(tprm_prm_link.id),
    #             '{"a": 1}',
    #             [1, 2, 3],
    #             ["abc"],
    #             [1.2],
    #             ["2022-01-01"],
    #             ["2022-01-01 08:00:00"],
    #             ["true"],
    #             ["str-1"],
    #             [1],
    #         ],
    #         [
    #             "valid_string",
    #             "5",
    #             "1",
    #             "2023-05-10",
    #             "2023-05-10 10:00:00",
    #             str(tprm_prm_link.id),
    #             '{"a": 1}',
    #             [100],
    #             ["abc", "def", "ghi"],
    #             [1.1, 2.2, 3.3],
    #             ["2023-01-01"],
    #             None,
    #             None,
    #             ["str-1", "str-2"],
    #             [1, 2],
    #         ],
    #         [
    #             "valid_string",
    #             "5",
    #             "0",
    #             "2023-05-10",
    #             "2023-05-10 10:00:00",
    #             str(tprm_prm_link.id),
    #             '{"a": 1}',
    #             [1, 2, 3, 4],
    #             ["a", "b"],
    #             [0.1],
    #             ["2022-01-01"],
    #             ["2022-01-01 08:15:00"],
    #             ["false"],
    #             ["str-5"],
    #             [5],
    #         ],
    #         [
    #             "valid_string",
    #             "5",
    #             "0",
    #             "2023-05-10",
    #             "2023-05-10 10:00:00",
    #             str(tprm_prm_link.id),
    #             '{"a": 1}',
    #             [3],
    #             ["def"],
    #             [0.2],
    #             ["2023-02-01"],
    #             ["2023-02-01 12:00:00"],
    #             ["false"],
    #             None,
    #             [2, 3],
    #         ],
    #         [
    #             "second_valid_string",
    #             "10",
    #             "0",
    #             "2002-12-12",
    #             "2002-12-12 18:45:30",
    #             str(tprm_prm_link.id),
    #             '{"a": 1}',
    #             [3, 5],
    #             ["ghi"],
    #             [1.3],
    #             ["2024-01-01"],
    #             ["2024-01-01 14:30:00"],
    #             ["true"],
    #             ["str-3"],
    #             [3, 4],
    #         ],
    #         [
    #             "third_valid_string",
    #             "193403",
    #             "true",
    #             "2022-01-01",
    #             "2022-01-01 08:15:00",
    #             str(tprm_prm_link.id),
    #             '{"a": 1}',
    #             [100],
    #             ["xyz"],
    #             [2.0],
    #             ["2023-03-01"],
    #             ["2023-03-01 10:00:00"],
    #             ["false"],
    #             ["str-2"],
    #             [6, 7],
    #         ],
    #     ]
    #
    #     file = generate_csv_in_memory(file_data)
    #
    #     response = client.post(
    #         URL + f"{tmo.id}",
    #         data={"filename": file.name, "type": "multipart/form-data"},
    #         files={"file": file},
    #     )
    #
    #     assert response.status_code == 200
    #     response_stream = BytesIO(response.content)
    #     response_stream.seek(0)
    #
    #     workbook = openpyxl.load_workbook(response_stream)
    #     # workbook.save('workbook.xlsx')
    #
    #     actual_types = list()
    #     data_model_sheet = workbook["Data Model"]
    #     max_row = data_model_sheet.max_row
    #
    #     for row in data_model_sheet.iter_rows(
    #         min_row=2,
    #         max_row=max_row,
    #         values_only=True,
    #     ):
    #         (
    #             tprm_name,
    #             file_column_type,
    #             inventory_column_name,
    #             inventory_column_type,
    #             _,
    #             _,
    #             _,
    #         ) = row
    #         actual_types.append(
    #             [
    #                 tprm_name,
    #                 file_column_type,
    #                 inventory_column_name,
    #                 inventory_column_type,
    #             ]
    #         )


#
#     expected_types = [
#         ["Test str TPRM", "str", "Test str TPRM", "str"],
#         ["Test int TPRM", "int", "Test int TPRM", "int"],
#         ["Test bool TPRM", "int", "Test bool TPRM", "bool"],
#         ["Test date TPRM", "str", "Test date TPRM", "date"],
#         ["Test datetime TPRM", "str", "Test datetime TPRM", "datetime"],
#         ["Test prm_link TPRM", "int", "Test prm_link TPRM", "prm_link"],
#         ["geometry", "json", "geometry", "json"],
#         ["Test multiple int TPRM", "list", "Test multiple int TPRM", "int"],
#         ["Test multiple str TPRM", "list", "Test multiple str TPRM", "str"],
#         [
#             "Test multiple float TPRM",
#             "list",
#             "Test multiple float TPRM",
#             "float",
#         ],
#         ["Test multiple date TPRM", "list", "Test multiple date TPRM", "date"],
#         [
#             "Test multiple datetime TPRM",
#             "list",
#             "Test multiple datetime TPRM",
#             "datetime",
#         ],
#         ["Test multiple bool TPRM", "list", "Test multiple bool TPRM", "bool"],
#         [
#             "Test multiple mo_link TPRM",
#             "list",
#             "Test multiple mo_link TPRM",
#             "mo_link",
#         ],
#         [
#             "Test multiple prm_link TPRM",
#             "list",
#             "Test multiple prm_link TPRM",
#             "prm_link",
#         ],
#     ]
#
#     assert actual_types == expected_types
#
#
# def test_validate_full_report_sheets(session: Session, client: TestClient):
#     """The test mainly devoted to report sheets creation."""
#     tmo = session.exec(select(TMO).where(TMO.name == "Test TMO")).first()
#
#     file_data = INPUT_SHEET_DATA_EXAMPLE
#
#     file = generate_csv_in_memory(file_data)
#
#     response = client.post(
#         URL + f"{tmo.id}",
#         data={"filename": file.name, "type": "multipart/form-data"},
#         files={"file": file},
#     )
#
#     assert response.status_code == 200
#     response_stream = BytesIO(response.content)
#     response_stream.seek(0)
#
#     workbook = openpyxl.load_workbook(response_stream)
#     # workbook.save('workbook.xlsx')
#
#     expected_sheetnames = {"Summary", "Data Model", "Update", "New", "Errors"}
#
#     expected_summary_content = [
#         [
#             "Updated objects",
#             "Created objects",
#             "Updated parameters",
#             "Created parameters",
#             "Deleted parameters",
#         ],
#         [0, 8, 0, 86, 0],
#     ]
#
#     expected_data_model_content = [
#         [
#             "File column name",
#             "File column type",
#             "Inventory column name",
#             "Inventory column type",
#             "Constraint",
#             "Required",
#             "Status",
#         ],
#         ["Test str TPRM", "str", "Test str TPRM", "str", None, True, "OK"],
#         ["Test int TPRM", "int", "Test int TPRM", "int", None, True, "ERROR"],
#         [
#             "Test bool TPRM",
#             "int",
#             "Test bool TPRM",
#             "bool",
#             None,
#             False,
#             "ERROR",
#         ],
#         [
#             "Test date TPRM",
#             "str",
#             "Test date TPRM",
#             "date",
#             None,
#             False,
#             "ERROR",
#         ],
#         [
#             "Test datetime TPRM",
#             "str",
#             "Test datetime TPRM",
#             "datetime",
#             None,
#             False,
#             "ERROR",
#         ],
#         [
#             "Test prm_link TPRM",
#             "int",
#             "Test prm_link TPRM",
#             "prm_link",
#             'Parameter type "Test int TPRM"',
#             False,
#             "ERROR",
#         ],
#         ["geometry", "json", "geometry", "json", None, None, "OK"],
#         [
#             "Test multiple int TPRM",
#             "list",
#             "Test multiple int TPRM",
#             "int",
#             None,
#             False,
#             "WARNING",
#         ],
#         [
#             "Test multiple str TPRM",
#             "list",
#             "Test multiple str TPRM",
#             "str",
#             "^[a-zA-Z]+$",
#             False,
#             "OK",
#         ],
#         [
#             "Test multiple float TPRM",
#             "list",
#             "Test multiple float TPRM",
#             "float",
#             None,
#             False,
#             "OK",
#         ],
#         [
#             "Test multiple date TPRM",
#             "list",
#             "Test multiple date TPRM",
#             "date",
#             None,
#             False,
#             "ERROR",
#         ],
#         [
#             "Test multiple datetime TPRM",
#             "list",
#             "Test multiple datetime TPRM",
#             "datetime",
#             None,
#             False,
#             "OK",
#         ],
#         [
#             "Test multiple bool TPRM",
#             "list",
#             "Test multiple bool TPRM",
#             "bool",
#             None,
#             False,
#             "WARNING",
#         ],
#         [
#             "Test multiple mo_link TPRM",
#             "list",
#             "Test multiple mo_link TPRM",
#             "mo_link",
#             None,
#             False,
#             "ERROR",
#         ],
#         [
#             "Test multiple prm_link TPRM",
#             "list",
#             "Test multiple prm_link TPRM",
#             "prm_link",
#             'Parameter type "Test str TPRM"',
#             False,
#             "ERROR",
#         ],
#     ]
#
#     expected_new_content = [
#         [
#             "Parent Name",
#             "Object Name",
#             "Test str TPRM",
#             "Test int TPRM",
#             "Test bool TPRM",
#             "Test date TPRM",
#             "Test datetime TPRM",
#             "Test prm_link TPRM",
#             "geometry",
#             "Test multiple int TPRM",
#             "Test multiple str TPRM",
#             "Test multiple float TPRM",
#             "Test multiple date TPRM",
#             "Test multiple datetime TPRM",
#             "Test multiple bool TPRM",
#             "Test multiple mo_link TPRM",
#             "Test multiple prm_link TPRM",
#         ],
#         [
#             None,
#             "fifth_valid_string-193403",
#             "fifth_valid_string",
#             193403,
#             True,
#             "2022-01-01",
#             "2022-01-01 12:30:45",
#             None,
#             '{"a": 1}',
#             "[23434, 123]",
#             '["ABC", "ABC", "ABC", "ABC", "ABC", "ABC", "ABC", "ABC", "ABC", "ABC"]',
#             "[1.2, 2.3]",
#             '["2021-01-01", "2022-02-02"]',
#             None,
#             "[true, true, true]",
#             '["str-1", "str-2"]',
#             "[]",
#         ],
#         [
#             None,
#             "first_valid_string-5",
#             "first_valid_string",
#             5,
#             True,
#             "2023-05-10",
#             "2023-05-10 10:00:00",
#             None,
#             '{"a": 1}',
#             "[1, 2, 3]",
#             '["abc"]',
#             "[1.2]",
#             '["2022-01-01"]',
#             '["2022-01-01 08:00:00"]',
#             "[true]",
#             '["str-1"]',
#             '["mo 1 value"]',
#         ],
#         [
#             None,
#             "fourth_invalid_string-4354",
#             "fourth_invalid_string",
#             4354,
#             True,
#             None,
#             None,
#             "3",
#             '{"a": 1}',
#             "[123, 20, 20]",
#             '["A", "B", "C"]',
#             "[1.1, 2.2, 3.3]",
#             '["2021-01-01", "2021-01-01", "2021-01-01", "2021-01-01", "2021-01-01", "2021-01-01", "2021-01-01",'
#             ' "2021-01-01", "2021-01-01", "2021-01-01"]',
#             None,
#             "[true, false]",
#             '["str-1", "str-1"]',
#             None,
#         ],
#         [
#             None,
#             "second_valid_string-10",
#             "second_valid_string",
#             10,
#             False,
#             "2002-12-12",
#             "2002-12-12 18:45:30",
#             None,
#             '{"a": 1}',
#             "[3, 5]",
#             '["ghi"]',
#             "[1.3]",
#             None,
#             '["2024-01-01 14:30:00"]',
#             "[true]",
#             '["parent_object-str-3"]',
#             None,
#         ],
#         [
#             None,
#             "sixth_valid_string-193403",
#             "sixth_valid_string",
#             193403,
#             True,
#             "2022-01-01",
#             None,
#             None,
#             '{"a": 1}',
#             "[1, 0, 1]",
#             '["abc", "def", "ghi"]',
#             "[0.1]",
#             '["2018-08-08"]',
#             None,
#             "[true, false, false, false]",
#             None,
#             '["mo 1 value", "str"]',
#         ],
#         [
#             None,
#             "third_invalid_string-12",
#             "third_invalid_string",
#             12,
#             True,
#             None,
#             None,
#             None,
#             '{"a": 1}',
#             "[12, -20, 10]",
#             '["ABrwrW"]',
#             "[1.0, 2.0, 3.0, 4.0, 5.0]",
#             None,
#             '["2022-01-01 00:00:00"]',
#             "[true, false, true, false]",
#             None,
#             "[]",
#         ],
#         [
#             None,
#             "third_valid_string-193403",
#             "third_valid_string",
#             193403,
#             True,
#             "2022-01-01",
#             "2022-01-01 08:15:00",
#             None,
#             '{"a": 1}',
#             "[100]",
#             '["xyz"]',
#             "[2.0]",
#             '["2023-03-01"]',
#             '["2023-03-01 10:00:00"]',
#             "[false]",
#             '["str-2"]',
#             None,
#         ],
#         [
#             None,
#             "valid_string-5",
#             "valid_string",
#             5,
#             True,
#             "2023-05-10",
#             "2023-05-10 10:00:00",
#             None,
#             '{"a": 1}',
#             "[100]",
#             '["abc", "def", "ghi"]',
#             "[1.1, 2.2, 3.3]",
#             '["2023-01-01"]',
#             None,
#             None,
#             '["str-1", "str-2"]',
#             '["mo 1 value", "str"]',
#         ],
#     ]
#
#     expected_update_content = [
#         [
#             "Parent Name",
#             "Object Name",
#             "Parameter Name",
#             "Old Value",
#             "New Value",
#         ]
#     ]
#
#     actual_sheetnames = set(workbook.sheetnames)
#
#     assert expected_sheetnames.difference(actual_sheetnames) == set()
#
#     summary_sheet = workbook["Summary"]
#     summary_max_row = summary_sheet.max_row
#     actual_summary_content = list()
#
#     for row in summary_sheet.iter_rows(
#         min_row=1,
#         max_row=summary_max_row,
#         values_only=True,
#     ):
#         actual_summary_content.append(list(row))
#
#     assert actual_summary_content == expected_summary_content
#
#     data_model_sheet = workbook["Data Model"]
#     data_model_max_row = data_model_sheet.max_row
#     actual_data_model_content = list()
#     for row in data_model_sheet.iter_rows(
#         min_row=1,
#         max_row=data_model_max_row,
#         values_only=True,
#     ):
#         actual_data_model_content.append(list(row))
#
#     assert actual_data_model_content == expected_data_model_content
#
#     update_sheet = workbook["Update"]
#     update_max_row = update_sheet.max_row
#     actual_update_content = list()
#     for row in update_sheet.iter_rows(
#         min_row=1,
#         max_row=update_max_row,
#         values_only=True,
#     ):
#         actual_update_content.append(list(row))
#
#     assert actual_update_content == expected_update_content
#
#     new_sheet = workbook["New"]
#     new_max_row = new_sheet.max_row
#     actual_new_content = list()
#     for row in new_sheet.iter_rows(
#         min_row=1, max_row=new_max_row, values_only=True
#     ):
#         actual_new_content.append(list(row))
#
#     assert actual_new_content == expected_new_content
#
#     errors_sheet = workbook["Errors"]
#     errors_max_row = errors_sheet.max_row
#     actual_errors_content = list()
#     for row in errors_sheet.iter_rows(
#         min_row=1, max_row=errors_max_row, values_only=True
#     ):
#         actual_errors_content.append(list(row))
#     expected_errors_content = [
#         [
#             "Parent Name",
#             "Object Name",
#             "Parameter Name",
#             "Value",
#             "Reason",
#             "Row ID",
#         ],
#         [
#             None,
#             "fifth_valid_string-193403",
#             "Test prm_link TPRM",
#             "2",
#             "You try to add not exists object",
#             5,
#         ],
#         [
#             None,
#             "first_valid_string-5",
#             "Test prm_link TPRM",
#             "14",
#             "You try to add not exists object",
#             7,
#         ],
#         [
#             None,
#             "fourth_invalid_string-4354",
#             "Test date TPRM",
#             "xbnepgijec",
#             "Every TPRM has own value type, but this value doesn't match it",
#             4,
#         ],
#         [
#             None,
#             "fourth_invalid_string-4354",
#             "Test datetime TPRM",
#             "2022-01-01 14:75:00",
#             "Every TPRM has own value type, but this value doesn't match it",
#             4,
#         ],
#         [
#             None,
#             "second_valid_string-10",
#             "Test prm_link TPRM",
#             "14",
#             "You try to add not exists object",
#             11,
#         ],
#         [
#             None,
#             "second_valid_string-10",
#             "Test multiple date TPRM",
#             "['24-01-01']",
#             "Every TPRM has own value type, but this value doesn't match it",
#             11,
#         ],
#         [
#             None,
#             "second_valid_string-10",
#             "Test multiple prm_link TPRM",
#             "[3, 4]",
#             "You try to add not exists object",
#             11,
#         ],
#         [
#             None,
#             "sixth_valid_string-193403",
#             "Test datetime TPRM",
#             "2022",
#             "Every TPRM has own value type, but this value doesn't match it",
#             6,
#         ],
#         [
#             None,
#             "sixth_valid_string-193403",
#             "Test prm_link TPRM",
#             "ertret",
#             "Every TPRM has own value type, but this value doesn't match it",
#             6,
#         ],
#         [
#             None,
#             "third_invalid_string-12",
#             "Test date TPRM",
#             "2020-45-02",
#             "Every TPRM has own value type, but this value doesn't match it",
#             3,
#         ],
#         [
#             None,
#             "third_invalid_string-12",
#             "Test datetime TPRM",
#             "2020-12-30 25:61:00",
#             "Every TPRM has own value type, but this value doesn't match it",
#             3,
#         ],
#         [
#             None,
#             "third_invalid_string-12",
#             "Test prm_link TPRM",
#             "2",
#             "You try to add not exists object",
#             3,
#         ],
#         [
#             None,
#             "third_valid_string-193403",
#             "Test prm_link TPRM",
#             "14",
#             "You try to add not exists object",
#             12,
#         ],
#         [
#             None,
#             "third_valid_string-193403",
#             "Test multiple prm_link TPRM",
#             "[6, 7]",
#             "You try to add not exists object",
#             12,
#         ],
#         [
#             None,
#             "valid_string-5",
#             "Test prm_link TPRM",
#             "14",
#             "You try to add not exists object",
#             8,
#         ],
#         [
#             None,
#             "valid_string-5",
#             "Object Name",
#             "valid_string-5",
#             "This object name appears more than once at indexes: [7, 8, 9]",
#             8,
#         ],
#         [
#             None,
#             None,
#             "Test int TPRM",
#             "invalid_int_value",
#             "Every TPRM has own value type, but this value doesn't match it",
#             1,
#         ],
#         [
#             None,
#             None,
#             "Test bool TPRM",
#             "WROYEUW",
#             "Every TPRM has own value type, but this value doesn't match it",
#             2,
#         ],
#         [
#             None,
#             None,
#             "Test datetime TPRM",
#             "123",
#             "Every TPRM has own value type, but this value doesn't match it",
#             2,
#         ],
#         [
#             None,
#             None,
#             "Test prm_link TPRM",
#             "1",
#             "You try to add not exists object",
#             2,
#         ],
#         [
#             None,
#             None,
#             "Test prm_link TPRM",
#             "14",
#             "You try to add not exists object",
#             9,
#         ],
#         [
#             None,
#             None,
#             "Test multiple mo_link TPRM",
#             "['str-1', 'str-2', '\"parent_object-str-3\"']",
#             "You try to add not exists object",
#             2,
#         ],
#         [
#             None,
#             None,
#             "Test multiple mo_link TPRM",
#             "['str-5']",
#             "You try to add not exists object",
#             9,
#         ],
#         [
#             None,
#             None,
#             "Test multiple prm_link TPRM",
#             "[2, 3]",
#             "You try to add not exists object",
#             10,
#         ],
#     ]
#
#     assert actual_errors_content == expected_errors_content
#
#     # The 'Error' sheet was thoroughly tested in the other tests


@pytest.mark.skip(reason="Too big test 2")
def test_check_batch_preview_speed(session: Session, client: TestClient):
    values_to_create = 10_000

    for parent_object_id in range(0, values_to_create):
        parent_object = MO(
            name=f"parent_object-{parent_object_id}", tmo_id=1, active=True
        )
        session.add(parent_object)
    session.commit()
    data_for_file = []

    for index in range(0, values_to_create):
        data_for_file.append(
            [
                f"some_string-{index}",
                index,
                {f"new_geo{index}": index},
                str(index),
            ]
        )

    res = session.exec(select(TMO).where(TMO.id == 2)).first()
    file_data_with_parents = [
        ["Test str TPRM", "Test int TPRM", "geometry", "p_id"],
    ]
    file_data_with_parents.extend(data_for_file)

    file = generate_csv_in_memory(file_data_with_parents)
    start = datetime.now()
    res = client.post(
        URL + f"{res.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    print(datetime.now() - start)
    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    sheet = workbook["New"]

    header = [cell.value for cell in sheet[1]]
    assert sheet.max_row == values_to_create + 1  # +1 because of header
    assert header == [
        "Parent Name",
        "Object Name",
        "Test str TPRM",
        "Test int TPRM",
        "geometry",
    ]


def test_mo_link_primary_proces(session: Session, client: TestClient):
    parent_tmo = TMO(
        **{
            "name": "Parent TMO",
            "version": 1,
            "created_by": "Test admin",
            "modified_by": "Test admin",
        }
    )
    session.add(parent_tmo)

    session.commit()

    parent_tmo = session.execute(
        select(TMO).where(TMO.name == "Parent TMO")
    ).scalar()
    main_tmo = TMO(
        **{
            "name": "Main TMO",
            "version": 1,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "global_uniqueness": False,
            "p_id": parent_tmo.id,
        }
    )
    session.add(main_tmo)
    session.commit()

    mo_0 = MO(name="parent_object_0", tmo_id=parent_tmo.id, active=True)
    mo_1 = MO(name="parent_object_1", tmo_id=parent_tmo.id, active=True)
    mo_2 = MO(name="parent_object_2", tmo_id=parent_tmo.id, active=True)
    mo_3 = MO(name="parent_object_3", tmo_id=parent_tmo.id, active=True)
    session.add_all([mo_0, mo_1, mo_2, mo_3])

    main_tmo = session.execute(
        select(TMO).where(TMO.name == "Main TMO")
    ).scalar()

    TPRM_STR_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_primary_str_TPRM",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_MO_LINK_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_primary_mo_link_TPRM",
            "val_type": "mo_link",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    session.add_all([TPRM_STR_DEFAULT_DATA, TPRM_MO_LINK_DEFAULT_DATA])

    session.commit()

    str_tprm_id = (
        session.execute(
            select(TPRM).where(TPRM.name == "Test_primary_str_TPRM")
        )
        .scalar()
        .id
    )

    mo_link_tprm_id = (
        session.execute(
            select(TPRM).where(TPRM.name == "Test_primary_mo_link_TPRM")
        )
        .scalar()
        .id
    )

    main_tmo.primary = [str_tprm_id, mo_link_tprm_id]
    session.add(main_tmo)
    session.commit()
    session.flush()

    main_tmo = session.execute(
        select(TMO).where(TMO.name == "Main TMO")
    ).scalar()

    file_data_with_parents = [
        ["Test_primary_str_TPRM", "Test_primary_mo_link_TPRM", "parent_name"],
        ["some_string_1", "parent_object_0", "parent_object_0"],
        ["some_string_2", "parent_object_1", "parent_object_1"],
        ["some_string_3", "parent_object_2", "parent_object_2"],
        ["some_string_4", "parent_object_3", "parent_object_3"],
    ]

    file = generate_csv_in_memory(file_data_with_parents)
    res = client.post(
        URL + f"{main_tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    if res.status_code != 200:
        print(res.json())

    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    new_sheet = workbook["New"]

    # expected headers: ['Object name', 'Parent name', 'Test str TPRM', 'Test int TPRM', 'geometry']
    expected_rows = [
        [
            "Parent Name",
            "Object Name",
            "Test_primary_str_TPRM",
            "Test_primary_mo_link_TPRM",
        ],
        [
            "parent_object_0",
            "parent_object_0-some_string_1-parent_object_0",
            "some_string_1",
            "parent_object_0",
        ],
        [
            "parent_object_1",
            "parent_object_1-some_string_2-parent_object_1",
            "some_string_2",
            "parent_object_1",
        ],
        [
            "parent_object_2",
            "parent_object_2-some_string_3-parent_object_2",
            "some_string_3",
            "parent_object_2",
        ],
        [
            "parent_object_3",
            "parent_object_3-some_string_4-parent_object_3",
            "some_string_4",
            "parent_object_3",
        ],
    ]
    actual_new_rows = list()

    for row in new_sheet.iter_rows(
        min_row=1, max_row=new_sheet.max_row, values_only=True
    ):
        actual_new_rows.append(list(row))

    assert expected_rows == actual_new_rows


def test_latitude_and_longitude_values(session: Session, client: TestClient):
    parent_tmo = TMO(
        **{
            "name": "Parent TMO",
            "version": 1,
            "created_by": "Test admin",
            "modified_by": "Test admin",
        }
    )
    session.add(parent_tmo)

    session.commit()

    parent_tmo = session.execute(
        select(TMO).where(TMO.name == "Parent TMO")
    ).scalar()
    main_tmo = TMO(
        **{
            "name": "Main TMO",
            "version": 1,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "global_uniqueness": False,
            "p_id": parent_tmo.id,
        }
    )
    session.add(main_tmo)
    session.commit()

    mo_0 = MO(name="parent_object_0", tmo_id=parent_tmo.id, active=True)
    mo_1 = MO(name="parent_object_1", tmo_id=parent_tmo.id, active=True)
    mo_2 = MO(name="parent_object_2", tmo_id=parent_tmo.id, active=True)
    mo_3 = MO(name="parent_object_3", tmo_id=parent_tmo.id, active=True)
    session.add_all([mo_0, mo_1, mo_2, mo_3])

    main_tmo = session.execute(
        select(TMO).where(TMO.name == "Main TMO")
    ).scalar()

    TPRM_LAT_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_latitude",
            "val_type": "float",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_LONG_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_longitude",
            "val_type": "float",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_STATUS_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_status",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_PRIMARY_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_primary",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    session.add_all(
        [
            TPRM_LAT_DEFAULT_DATA,
            TPRM_LONG_DEFAULT_DATA,
            TPRM_STATUS_DEFAULT_DATA,
            TPRM_PRIMARY_DEFAULT_DATA,
        ]
    )

    session.commit()

    lat_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_latitude"))
        .scalar()
        .id
    )

    long_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_longitude"))
        .scalar()
        .id
    )

    status_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_status"))
        .scalar()
        .id
    )

    primary_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_primary"))
        .scalar()
        .id
    )

    main_tmo.latitude = lat_tprm_id
    main_tmo.longitude = long_tprm_id
    main_tmo.status = status_tprm_id
    main_tmo.primary = [primary_tprm_id]
    session.add(main_tmo)
    session.commit()
    session.flush()

    main_tmo = session.execute(
        select(TMO).where(TMO.name == "Main TMO")
    ).scalar()

    already_exists_object = MO(
        name="parent_object_0-already_exists_object",
        tmo_id=main_tmo.id,
        active=True,
    )
    session.add(already_exists_object)
    session.commit()

    file_data_with_parents = [
        [
            "Test_latitude",
            "Test_longitude",
            "parent_name",
            "Test_status",
            "Test_primary",
        ],
        [
            "42.12314",
            "73.3453453",
            "parent_object_0",
            "status1",
            "already_exists_object",
        ],
        ["43.2342526", "52", "parent_object_1", "status2", "new_1"],
        [
            "52.234234123124",
            "some_str_value",
            "parent_object_2",
            "status3",
            "new_2",
        ],
        ["12.121435", "54.23623523", "parent_object_3", "status4", "new_3"],
    ]

    file = generate_csv_in_memory(file_data_with_parents)
    res = client.post(
        URL + f"{main_tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    if res.status_code != 200:
        print(res.json())

    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    new_sheet = workbook["New"]

    expected_rows = [
        [
            "Parent Name",
            "Object Name",
            "Test_latitude",
            "Test_longitude",
            "Test_status",
            "Test_primary",
            "latitude",
            "longitude",
            "status",
        ],
        [
            "parent_object_1",
            "parent_object_1-new_1",
            43.2342526,
            52,
            "status2",
            "new_1",
            43.2342526,
            52,
            "status2",
        ],
        [
            "parent_object_2",
            "parent_object_2-new_2",
            52.234234123124,
            None,
            "status3",
            "new_2",
            52.234234123124,
            None,
            "status3",
        ],
        [
            "parent_object_3",
            "parent_object_3-new_3",
            12.121435,
            54.23623523,
            "status4",
            "new_3",
            12.121435,
            54.23623523,
            "status4",
        ],
    ]

    actual_new_rows = list()

    for row in new_sheet.iter_rows(
        min_row=1, max_row=new_sheet.max_row, values_only=True
    ):
        actual_new_rows.append(list(row))

    assert expected_rows == actual_new_rows

    update_sheet = workbook["Update"]

    expected_rows = [
        [
            "Parent Name",
            "Object Name",
            "Parameter Name",
            "Old Value",
            "New Value",
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "Test_latitude",
            None,
            42.12314,
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "Test_longitude",
            None,
            73.3453453,
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "Test_status",
            None,
            "status1",
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "parent_name",
            None,
            "parent_object_0",
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "latitude",
            None,
            42.12314,
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "longitude",
            None,
            73.3453453,
        ],
        [
            "parent_object_0",
            "parent_object_0-already_exists_object",
            "status",
            None,
            "status1",
        ],
    ]
    actual_update_rows = list()

    for row in update_sheet.iter_rows(
        min_row=1, max_row=update_sheet.max_row, values_only=True
    ):
        actual_update_rows.append(list(row))

    assert expected_rows == actual_update_rows

    error_sheet = workbook["Errors"]

    expected_rows = [
        [
            "Parent Name",
            "Object Name",
            "Parameter Name",
            "Value",
            "Reason",
            "Row ID",
        ],
        [
            "parent_object_2",
            "parent_object_2-new_2",
            "Test_longitude",
            "some_str_value",
            "Every TPRM has own value type, but this value doesn't match it",
            3,
        ],
    ]

    actual_error_rows = list()

    for row in error_sheet.iter_rows(
        min_row=1, max_row=error_sheet.max_row, values_only=True
    ):
        actual_error_rows.append(list(row))

    assert actual_error_rows == expected_rows


@pytest.mark.skip(reason="Too big test")
def test_global_test(session: Session, client: TestClient):
    LINE_COUNT = 100_000

    parent_tmo = TMO(
        **{
            "name": "Parent TMO",
            "version": 1,
            "created_by": "Test admin",
            "modified_by": "Test admin",
        }
    )
    session.add(parent_tmo)

    session.commit()

    parent_tmo = session.execute(
        select(TMO).where(TMO.name == "Parent TMO")
    ).scalar()
    main_tmo = TMO(
        **{
            "name": "Main TMO",
            "version": 1,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "global_uniqueness": False,
            "p_id": parent_tmo.id,
        }
    )
    session.add(main_tmo)
    session.commit()

    PARENT_TPRM_STR_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_parent_str",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    PARENT_TPRM_MULTIPLE_STR_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_parent_str_mult",
            "val_type": "str",
            "required": True,
            "multiple": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    mo_0 = MO(name="parent_object_0", tmo_id=parent_tmo.id, active=True)
    mo_1 = MO(name="parent_object_1", tmo_id=parent_tmo.id, active=True)
    mo_2 = MO(name="parent_object_2", tmo_id=parent_tmo.id, active=True)
    mo_3 = MO(name="parent_object_3", tmo_id=parent_tmo.id, active=True)
    session.add_all(
        [
            mo_0,
            mo_1,
            mo_2,
            mo_3,
            PARENT_TPRM_STR_DEFAULT_DATA,
            PARENT_TPRM_MULTIPLE_STR_DEFAULT_DATA,
        ]
    )
    session.commit()

    parent_tprm_str = session.execute(
        select(TPRM).where(TPRM.name == "Test_parent_str")
    ).scalar()
    parent_tprm_multiple_str = session.execute(
        select(TPRM).where(TPRM.name == "Test_parent_str_mult")
    ).scalar()

    mo_0 = session.execute(
        select(MO).where(MO.name == "parent_object_0")
    ).scalar()  # 8
    mo_1 = session.execute(
        select(MO).where(MO.name == "parent_object_1")
    ).scalar()  # 9
    mo_2 = session.execute(
        select(MO).where(MO.name == "parent_object_2")
    ).scalar()  # 10
    mo_3 = session.execute(
        select(MO).where(MO.name == "parent_object_3")
    ).scalar()  # 11

    prm_0 = PRM(value="some_value_1", tprm_id=parent_tprm_str.id, mo_id=mo_0.id)
    prm_1 = PRM(value="some_value_2", tprm_id=parent_tprm_str.id, mo_id=mo_1.id)
    prm_2 = PRM(
        value=pickle.dumps(["some_value_3", "some_value_4"]).hex(),
        tprm_id=parent_tprm_multiple_str.id,
        mo_id=mo_2.id,
    )
    prm_3 = PRM(
        value=pickle.dumps(["some_value_5", "some_value_6"]).hex(),
        tprm_id=parent_tprm_multiple_str.id,
        mo_id=mo_3.id,
    )
    session.add_all([prm_0, prm_1, prm_2, prm_3])
    session.commit()

    prm_0 = session.execute(
        select(PRM).where(
            PRM.tprm_id == parent_tprm_str.id, PRM.mo_id == mo_0.id
        )
    ).scalar()  # 5
    prm_1 = session.execute(
        select(PRM).where(
            PRM.tprm_id == parent_tprm_str.id, PRM.mo_id == mo_1.id
        )
    ).scalar()  # 6
    prm_2 = session.execute(
        select(PRM).where(
            PRM.tprm_id == parent_tprm_multiple_str.id, PRM.mo_id == mo_2.id
        )
    ).scalar()  # 7
    prm_3 = session.execute(
        select(PRM).where(
            PRM.tprm_id == parent_tprm_multiple_str.id, PRM.mo_id == mo_3.id
        )
    ).scalar()  # 8

    main_tmo = session.execute(
        select(TMO).where(TMO.name == "Main TMO")
    ).scalar()

    TPRM_STR_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_str",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": "^(?=.*[A-Z])[A-Za-z0-9]+$",
        }
    )

    TPRM_INT_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_int",
            "val_type": "int",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_INT_MULTIPLE_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_int_mult",
            "val_type": "int",
            "multiple": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": "1:50000",
        }
    )

    TPRM_LAT_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_latitude",
            "val_type": "float",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": "1:50000",
        }
    )

    TPRM_LONG_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_longitude",
            "val_type": "float",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": "1:50000",
        }
    )

    TPRM_STATUS_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_status",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_PRIMARY_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_primary",
            "val_type": "str",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": "^(?=.*[A-Z])[A-Za-z0-9]+$",
        }
    )

    TPRM_MO_LINK_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_primary_mo_link",
            "val_type": "mo_link",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_PRM_LINK_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_prm_link",
            "val_type": "prm_link",
            "required": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": str(parent_tprm_str.id),
        }
    )

    TPRM_MO_LINK_MULTIPLE_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_mo_link_mult",
            "val_type": "mo_link",
            "required": True,
            "multiple": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
        }
    )

    TPRM_PRM_LINK_MULTIPLE_DEFAULT_DATA = TPRM(
        **{
            "name": "Test_prm_link_mult",
            "val_type": "prm_link",
            "required": True,
            "multiple": True,
            "created_by": "Test admin",
            "modified_by": "Test admin",
            "tmo_id": main_tmo.id,
            "constraint": str(parent_tprm_str.id),
        }
    )

    session.add_all(
        [
            TPRM_LAT_DEFAULT_DATA,
            TPRM_LONG_DEFAULT_DATA,
            TPRM_STATUS_DEFAULT_DATA,
            TPRM_PRIMARY_DEFAULT_DATA,
            TPRM_STR_DEFAULT_DATA,
            TPRM_INT_DEFAULT_DATA,
            TPRM_MO_LINK_DEFAULT_DATA,
            TPRM_PRM_LINK_DEFAULT_DATA,
            TPRM_INT_MULTIPLE_DEFAULT_DATA,
            TPRM_MO_LINK_MULTIPLE_DEFAULT_DATA,
            TPRM_PRM_LINK_MULTIPLE_DEFAULT_DATA,
        ]
    )

    session.commit()

    lat_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_latitude"))
        .scalar()
        .id
    )

    long_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_longitude"))
        .scalar()
        .id
    )

    status_tprm_id = (
        session.execute(select(TPRM).where(TPRM.name == "Test_status"))
        .scalar()
        .id
    )

    primary_tprm_id = (
        session.execute(
            select(TPRM.id).where(
                TPRM.name.in_(
                    [
                        "Test_str",
                        "Test_int",
                        "Test_primary",
                        "Test_primary_mo_link",
                    ]
                )
            )
        )
        .scalars()
        .all()
    )

    main_tmo.latitude = lat_tprm_id
    main_tmo.longitude = long_tprm_id
    main_tmo.status = status_tprm_id
    main_tmo.primary = primary_tprm_id
    session.add(main_tmo)
    session.commit()
    session.flush()

    main_tmo = session.execute(
        select(TMO).where(TMO.name == "Main TMO")
    ).scalar()

    already_exists_object = MO(
        name="parent_object_0-already_exists_object",
        tmo_id=main_tmo.id,
        active=True,
    )
    session.add(already_exists_object)
    session.commit()

    file_data_with_parents = [
        [
            "Test_str",
            "Test_int",
            "Test_int_mult",
            "Test_latitude",
            "Test_longitude",
            "Test_status",
            "Test_primary",
            "Test_primary_mo_link",
            "Test_prm_link",
            "Test_prm_link_mult",
            "Test_mo_link_mult",
            "parent_name",
        ],
    ]

    exists_prm_multiple = [[5, 6], [5], [6]]
    exists_parents_multiple = [
        ["parent_object_0", "parent_object_1"],
        ["parent_object_2", "parent_object_3"],
    ]
    exists_parents = [
        "parent_object_0",
        "parent_object_1",
        "parent_object_2",
        "parent_object_3",
    ]

    for index in range(0, LINE_COUNT):
        random.shuffle(exists_prm_multiple)
        random.shuffle(exists_parents)
        random.shuffle(exists_parents_multiple)

        test_str = f"PrimaryValue{index}"
        test_int = index
        test_int_mult = exists_prm_multiple[0]
        test_latitude = random.uniform(42.234252, 45.23423453)
        test_longitude = random.uniform(72.345235245, 92.1231252)
        test_status = f"SomeValue{index}"
        test_primary = f"SomeValue{index}"
        test_primary_mo_link = exists_parents[0]
        test_prm_link = int(random.uniform(5, 7))
        test_prm_link_mult = exists_prm_multiple[0]
        test_mo_link_mult = exists_parents_multiple[0]
        parent_name = exists_parents[0]

        file_data_with_parents.append(
            [
                test_str,
                test_int,
                test_int_mult,
                test_latitude,
                test_longitude,
                test_status,
                test_primary,
                test_primary_mo_link,
                test_prm_link,
                test_prm_link_mult,
                test_mo_link_mult,
                parent_name,
            ]
        )

    start = datetime.now()
    print("start", start)
    file = generate_csv_in_memory(file_data_with_parents)
    res = client.post(
        URL + f"{main_tmo.id}",
        data={"filename": file.name, "type": "multipart/form-data"},
        files={"file": file},
    )
    print("RESULT", datetime.now() - start)

    if res.status_code != 200:
        print(res.json())

    assert res.status_code == 200
    response_stream = BytesIO(res.content)
    response_stream.seek(0)

    workbook = openpyxl.load_workbook(response_stream)

    new_sheet = workbook["New"]

    expected_rows = [
        [
            "Object Name",
            "Parent Name",
            "Test_latitude",
            "Test_longitude",
            "Test_status",
            "Test_primary",
            "latitude",
            "longitude",
            "status",
        ],
        [
            "parent_object_1-new_1",
            "parent_object_1",
            43.2342526,
            52,
            "status2",
            "new_1",
            43.2342526,
            52,
            "status2",
        ],
        [
            "parent_object_2-new_2",
            "parent_object_2",
            52.234234123124,
            None,
            "status3",
            "new_2",
            52.234234123124,
            None,
            "status3",
        ],
        [
            "parent_object_3-new_3",
            "parent_object_3",
            12.121435,
            54.23623523,
            "status4",
            "new_3",
            12.121435,
            54.23623523,
            "status4",
        ],
    ]

    actual_new_rows = list()

    for row in new_sheet.iter_rows(
        min_row=1, max_row=new_sheet.max_row, values_only=True
    ):
        actual_new_rows.append(list(row))

    assert expected_rows == actual_new_rows

    update_sheet = workbook["Update"]

    expected_rows = [
        [
            "Parent Name",
            "Object Name",
            "Parameter Name",
            "Old Value",
            "New Value",
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "Test_latitude",
            None,
            42.12314,
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "Test_longitude",
            None,
            73.3453453,
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "Test_status",
            None,
            "status1",
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "parent_name",
            None,
            "parent_object_0",
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "latitude",
            None,
            42.12314,
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "longitude",
            None,
            73.3453453,
        ],
        [
            None,
            "parent_object_0-already_exists_object",
            "status",
            None,
            "status1",
        ],
    ]

    actual_update_rows = list()

    for row in update_sheet.iter_rows(
        min_row=1, max_row=update_sheet.max_row, values_only=True
    ):
        actual_update_rows.append(list(row))

    assert expected_rows == actual_update_rows
